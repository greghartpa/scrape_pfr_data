import os
import numpy as np
import pandas as pd
import requests
from bs4 import BeautifulSoup
from bs4 import Comment
import openpyxl
import time
import ssl

# Scrape player value data from PFR
url = 'https://www.pro-football-reference.com'

ssl._create_default_https_context = ssl._create_unverified_context
maxp = 10000
reccount = 0
totalcount = 0
skipped_count = 0
minyear = 0

# create dataframe to hold team abbreviations and correct teams that moved
data = {'OrigTm':['ARI','ATL','BAL','BUF','CAR',
    'CHI','CIN','CLE','DAL','DEN',
    'DET','GNB','HOU','IND','JAX',
    'KAN','LAC','LAR','LVR','MIA',
    'MIN','NOR','NWE','NYG','NYJ',
    'OAK','PHI','PHO','PIT','RAM',
    'SDG','SEA','SFO','STL','TAM',
    'TEN','WAS'],
    'Tm':['ARI','ATL','BAL','BUF','CAR',
    'CHI','CIN','CLE','DAL','DEN',
    'DET','GNB','HOU','IND','JAX',
    'KAN','LAC','LAR','LVR','MIA',
    'MIN','NOR','NWE','NYG','NYJ',
    'LVR','PHI','ARI','PIT','LAR',
    'LAC','SEA','SFO','LAR','TAM',
    'TEN','WFT']}
tmabrevdf = pd.DataFrame(data)
# Create dataframe of position groups to combine into the groupings used in draft data
data = {'Pos':['C','CB','DB','DE','DT','G','ILB','LB','NT','OL','OLB','QB','RB','S','T','TE','WR','K','P','LS','FB','DL','OG','OT',
        'FS','SS','EDGE','MLB','LG','RG','LILB','RILB','LOLB','ROLB','LT','RT','HB','LB-DE','DE-C','G-C','T-G','G-T','LB-DE',
        'LS-TE','QB-TE','WR/RB','WR/CB','DT/LB','PR-WR','NT-DT','DE-LB','DE-DT','DT-DE','G,C','QB/TE',
        'G,T','LS,TE','C-G','FB-LB','LLB','DT-NT','DT-DE-NT','C-T','TE-C','DE-DT-LB','C-TE',
        'T/G','T-TE','WR/QB','DE-NT-DT','LB/RB','C-T','DT-NT-LB','DE-T','RB/TE','G/C','LB-DB','RLB',

        'RCB', 'RDT', 'LCB', 'LDT', 'RDE', 'LDE', '/NT', 'DT/LDT', 'CB/LCB', 'RDE/RDT',
        'DE/LOLB', 'LT/T', '/TE', 'CB/DB', '/G', '/WR', 'DE/RDE', 'CB/RCB', 'LG/T', '/IL',
        'DT/LOLB', '/DE', 'G/RG', '/QB', 'DE/LDE', 'RT/T', 'IL/LILB', 'B', 'DE/RDT', '/DT',
        '/CB', '/C', '/FB', 'IL/RILB', '/SS', 'DB/FS', 'DT/RDE', 'IL/RLB', '/FS', 'DB/SS',
        '/T', 'DT/RDT', 'ML/MLB', 'LDT/NT', '/RB', 'G/LG', 'FS/SS', 'LILB/ML', 'G/TE',
        'C/LG', 'DT/LDE', 'G/LT', 'NT/RDT', 'C/RG', 'FB/TE', 'T,G', 'TE-FB', 'C,G',
        'TE-WR', 'FB-TE', 'LB/DE', 'G-C-T'],
        'PosGroup':['Off - IOL','Def - DB','Def - DB','Def - ED','Def - IDL','Off - IOL','Def - LB','Def - LB','Def - IDL',
        'Off - IOL','Def - LB','Off - QB','Off - RB','Def - S','Off - T','Off - TE','Off - WR','ST','ST','ST','Off - RB','Def - IDL',
        'Off - IOL','Off - T','Def - S','Def - S','Def - ED','Def - LB','Off - IOL','Off - IOL','Def - LB','Def - LB','Def - LB',
        'Def - LB','Off - T','Off - T','Off - RB','Def - LB','Def - ED','Off - IOL','Off - T','Off - T','Def - LB',
        'ST','Off - TE','Off - RB','Off - WR','Def - LB','Off - WR','Def - IDL','Def - LB','Def - IDL','Def - IDL','Off - IOL','Off - TE',
        'Off - IOL','Off - TE','Off - IOL','Def - LB','Def - LB','Def - IDL','Def - IDL','Off - IOL','Off - IOL','Def - ED','Off - IOL',
        'Off - IOL','Off - T','Off - QB','Def - IDL','Def - LB','Off - IOL','Def - IDL','Def - ED','Off - RB','Off - IOL','Def - LB','Def - LB',
        'Def - DB', 'Def - IDL', 'Def - DB', 'Def - IDL', 'Def - ED', 'Def - ED', 'Def - IDL', 'Def - IDL', 'Def - DB', 'Def - ED',
        'Def - LB', 'Off - T', 'Off - TE', 'Def - DB', 'Off - IOL', 'Off - WR', 'Def - ED', 'Def - DB', 'Off - IOL', 'Def - LB',
        'Def - IDL', 'Def - ED', 'Off - IOL', 'Off - QB', 'Def - ED', 'Off - T', 'Def - LB', 'Off - RB', 'Def - IDL', 'Def - IDL',
        'Def - DB', 'Off - IOL', 'Off - RB', 'Def - LB', 'Def - S', 'Def - S', 'Def - ED', 'Def - LB', 'Def - S', 'Def - S',
        'Off - T', 'Def - IDL', 'Def - LB', 'Def - IDL', 'Off - RB', 'Off - IOL', 'Def - S', 'Def - LB', 'Off - IOL',
        'Off - IOL', 'Def - ED', 'Off - IOL', 'Def - IDL', 'Off - IOL', 'Off - TE', 'Off - IOL', 'Off - TE', 'Off - IOL',
        'Off - TE', 'Off - TE', 'Def - LB', 'Off - IOL']}
posdf = pd.DataFrame(data)

# create data frame of position groupings as these will be used for more aggregated positional analysis
data = {'PosGroup':['Off - WR', 'Off - RB', 'Off - QB', 'Off - T', 'Off - TE', 'Off - OL', 'Off - IOL',
    'Def - IDL', 'Def - S', 'Def - DB', 'Def - ED', 'Def - LB', 'ST'],
    'PositionGroup':['Off - WR', 'Off - RB', 'Off - QB', 'Off - OL', 'Off - TE', 'Off - OL', 'Off - OL',
    'Def - DL', 'Def - DB', 'Def - DB', 'Def - DL', 'Def - LB', 'ST']}
positionsdf = pd.DataFrame(data)
column_names = ["Pos","PosGroup","PositionGroup"]
dfnewpos = pd.DataFrame(columns = column_names)

# create dataframe of starters
column_names = ["UniqueKey","Starter"]
starterdf = pd.DataFrame(columns = column_names)

# open the teams data xlsx to get roster urls and update status
teamsdatafilename = "teams-data.xlsx"
wb_obj = openpyxl.load_workbook(teamsdatafilename.strip())
sheet_obj = wb_obj.active
max_column=sheet_obj.max_column
max_row=sheet_obj.max_row

# read data from teams CSV
if os.path.isfile('player-data-inprocess.csv'):
    player_data_df = pd.read_csv("player-data-inprocess.csv")
    print("Found player data file - creating unique list")
    # uniqueplayers = player_data_df.Player.unique()
    # playersdf = pd.DataFrame(uniqueplayers, columns = ["Player"])
    uniqueplayers = player_data_df.PlayerKey.unique()
    playersdf = pd.DataFrame(uniqueplayers, columns = ["PlayerKey"])
else:
    print("Creating playersdf as does not exist already")
    column_names = ["PlayerKey"]
    playersdf = pd.DataFrame(columns = column_names)


# Loop through all team seasons
for index in range(2, max_row + 1):
    # create new dataframe here in loop as we will write it each time we finish a team season
    column_names = ["Yr", "Player", "PlayerKey", "Age", "Tm", "Pos", "G", "GS", "AV"]
    df = pd.DataFrame(columns = column_names)
    year = sheet_obj.cell(row=index,column=2).value
    team = sheet_obj.cell(row=index,column=3).value
    scrapedflag = sheet_obj.cell(row=index,column=23).value
    rosterurl = sheet_obj.cell(row=index,column=24).value

    if totalcount >= maxp:
        print('\nStopping - hit max players limit of ' + str(maxp))
        break

    # if the team has already been scraped, skip it to prevent duplicate data
    if scrapedflag == 0:
        if totalcount >= maxp:
            break
        print("Processing " + str(year) + " " + team)

        # grab this players stats
        maxretries = 10
        attempt = 0
        while attempt < maxretries:
            try:
                r = requests.get(rosterurl)
                # print(r)
            except Exception as e:
                print("IncompleteRead, retrying...")
                attempt += 1
                time.sleep(5)
            else:
                break
        soup = BeautifulSoup(r.content, 'html.parser')
        # Store list of starters so we can set the starter flag below
        startertable = soup.find(lambda tag: tag.name=='table' and tag.has_attr('id') and tag['id']=="starters")
        tempstarterdf = pd.read_html(str(startertable))[0]
        tempstarterdf = tempstarterdf[(tempstarterdf.Player != "Offensive Starters")]
        tempstarterdf = tempstarterdf[(tempstarterdf.Player != "Defensive Starters")]
        tempstarterdf["Yr"] = year
        tempstarterdf["OrigTm"] = team
        tempstarterdf["Starter"] = 1
        # Remove extra indictoars (pro bowl) from player name
        tempstarterdf["Player"] = tempstarterdf["Player"].map(lambda x: x.rstrip('*+'))
        tempstarterdf = pd.merge(tempstarterdf, tmabrevdf, on="OrigTm")
        tempstarterdf["UniqueKey"] = tempstarterdf["Yr"].astype(str) + "-" + tempstarterdf["Tm"] + "-" + tempstarterdf["Player"]
        # tempstarterdf = tempstarterdf[["Yr", "Tm", "Player", "UniqueKey", "Starter"]]
        tempstarterdf = tempstarterdf[["UniqueKey", "Starter"]]
        starterdf = starterdf.append(tempstarterdf, ignore_index=True)

        # Pull the comments as this is where the full roster is
        comments = soup.find_all(string=lambda text: isinstance(text, Comment))
        newsoup = BeautifulSoup(str(comments), 'html.parser')

        ptable = newsoup.find('table', attrs={'class':'per_match_toggle sortable stats_table'})
        ptable_body = ptable.find('tbody')
        for prow in ptable_body.find_all("tr"):
            if totalcount >= maxp:
                break

            # Get player data - name and link to data page
            for playerdata in prow.find_all("td", limit=1):

                # Get player name and their data page url
                player = playerdata.text
                playerlink = playerdata.find("a")
                playerlink = playerlink.get("href")
                playerurl = url + playerlink
                playerkey = playerurl[-12:]

                # Remove backslash escapes in the player name and URL as they will cause issues
                player = player.replace('\\', '')
                playerurl = playerurl.replace('\\', '')
                # Remove extra indictoars (pro bowl) from player name
                player = player.replace('*', '')
                player = player.replace('+', '')

                # don't process a player that's already been done - use the plyaerkey instead of player name
                # as players can have the same names
                if playersdf['PlayerKey'].str.contains(playerkey).any():
                    skipped_count = skipped_count + 1
                    #print("Skipping: " + player)
                else:

                    print("Scraping: " + player)
                    totalcount = totalcount + 1
                    # Add player to player name dataframe to ensure they aren't processed again
                    playersdf.loc[len(playersdf.index)] = [playerkey]
                    reccount = reccount + 1

                    # Need to get player position from top of page as yearly data is not reliable for position
                    maxretries = 10
                    attempt = 0
                    # print("first request " + playerurl)
                    while attempt < maxretries:
                        try:
                            tempr = requests.get(playerurl, verify=True) #, flavor="html5lib")
                        except Exception as e:
                            print("IncompleteRead 1, retrying...")
                            attempt += 1
                            time.sleep(5)
                        else:
                            break
                    tsoup = BeautifulSoup(tempr.content, 'html.parser')

                    alltags = tsoup.find_all("p", limit=2)
                    for tag in alltags:
                        startchar = tag.text.find("Position: ")
                        endchar = tag.text.find("Throws")
                        if startchar != -1:
                            if endchar != -1:
                                playerpos = tag.text[11:endchar].strip()
                            else:
                                playerpos = tag.text[11:].strip()

                    # grab this players stats
                    maxretries = 10
                    attempt = 0
                    while attempt < maxretries:
                        try:
                            pdf = pd.read_html(playerurl)[0]
                        #except http.client.IncompleteRead:
                        except Exception as e:
                            print("IncompleteRead 2, retrying...")
                            attempt += 1
                            time.sleep(5)
                        else:
                            break

                    # get rid of MultiIndex, just keep last row
                    pdf.columns = pdf.columns.get_level_values(-1)
                    # Add player name
                    pdf['Player'] = player
                    # Fix year by stripping pro bowl indicators
                    pdf['Year'] = pdf['Year'].astype(str)
                    pdf['Yr'] = pdf['Year'].str[:4]

                    # By default, use the top-level player position (pulled above from header) as the yearly records are
                    # often blank. But, if the yearly record has a position listed, use that position until another change is deteced
                    # as this identifies a position change by a player
                    # pdf.Pos.where((pdf.Pos == 'RLB'), playerpos, inplace=True)
                    pdf.Pos = pdf.Pos.fillna(playerpos)

                    # Add player URL key so we can uniquely match to players if they have the same name
                    pdf['PlayerKey'] = playerkey

                    # remove career summary line as we only want the individual years
                    pdf = pdf[pdf.Year != "Career"]
                    # prepare dataframe to be copied into end player dataframe
                    pdf = pdf[['Yr', 'Player', 'PlayerKey', 'Age', 'Tm', 'Pos', 'G', 'GS', 'AV']].copy()
                    pdf['Pos'] = pdf['Pos'].str.upper()
                    # Fix team abbreviation to handle team moves
                    pdf.rename(columns={"Tm": "OrigTm"}, inplace=True)
                    pdf = pd.merge(pdf, tmabrevdf, on="OrigTm")
                    pdf.drop(['OrigTm'], 1, inplace=True)

                    # Store in end dataframe to collect all players
                    df = df.append(pdf, ignore_index=True)

        # Remove blank records where full seasons were missed due to injury
        df = df[~df['Tm'].str.startswith('Missed')]
        # Remove career summary lines and partial season lines
        #df = df[df.Yr != "Care"]
        df = df[~df['Yr'].str.endswith('yr')]
        df = df[~df['Yr'].str.endswith('y')]

        # Fix year records that are split between multiple teams
        for ind in df.index:
            currentyr = df.at[ind, 'Yr']
            currentage = df.at[ind, 'Age']
            if (currentyr != "nan" and pd.isnull(currentyr) == False and currentyr != "*" and currentyr != "*+"):
                lastcapturedyr = currentyr
                lastcapturedage = currentage
            if (currentyr == "nan" or pd.isnull(currentyr) or currentyr == "*" or currentyr == "*+"):
                 df['Yr'][ind] = lastcapturedyr #df['Yr'][ind-1]
                 df['Age'][ind] = lastcapturedage #df['Age'][ind-1]

        # Remove records that summarize seasons split between teams as they were just fixed above
        df = df[~df['Tm'].str.endswith('TM')]

        # Convert numbers to actual number columns
        df["Age"] = pd.to_numeric(df["Age"])
        df["AV"] = pd.to_numeric(df["AV"])
        df["G"] = pd.to_numeric(df["G"])
        df["GS"] = pd.to_numeric(df["GS"])

        # Handle positions that are not captured above as they will be dropped
        temp = df.loc[~df['Pos'].isin(posdf['Pos'])].copy()
        temp['PosGroup'] = temp['Pos']
        temp['PositionGroup'] = temp['Pos']
        dfnewpos = temp[['Pos','PosGroup','PositionGroup']]
        dfnewpos = dfnewpos.drop_duplicates()
        temp = dfnewpos[['Pos','PosGroup']]
        posdf = posdf.append(temp)
        temp = dfnewpos[['PosGroup','PositionGroup']]
        positionsdf = positionsdf.append(temp)

        df = pd.merge(df, posdf, on="Pos", how="left")
        # Re-group positions into higher level groups for AV analysis
        df = pd.merge(df, positionsdf, on="PosGroup")

        # Calculate an AV per game in case needed to account for missed playing time
        df['AVperG'] = df['AV'] / df['G']
        df['AVperG'] = df['AVperG'].round(decimals=2)
        df['AVper16G'] = df['AVperG'] * 16
        df['AVper16G'] = df['AVper16G'].round(decimals=2)

        # Create adjusted AV to account for impact of nickel and dime defense on the calculations
        # AV assigns 2/3 of defensive team points to the front 7, 1/3 to the seconday
        # Current stats show NFL teams run "base" 30% of the time, nickel 50%, and dime 20%
        # Adjusting for the actual number of players on the field, you get the "front 7" should get
        # 87% of their current team points and teh secondary should get 122% of their current points
        df['AdjAV'] = df['AV']   # Start by setting Adjusted AV to AV as for most positions we are not touching it
        df.loc[df['PositionGroup'] == 'Def - LB', 'AdjAV'] = df['AV'] * 0.87
        df.loc[df['PositionGroup'] == 'Def - DL', 'AdjAV'] = df['AV'] * 0.87
        df.loc[df['PositionGroup'] == 'Def - DB', 'AdjAV'] = df['AV'] * 1.22
        df['AdjAVperG'] = df['AdjAV'] / df['G']
        df['AdjAVperG'] = df['AdjAVperG'].round(decimals=2)
        df['AdjAVper16G'] = df['AdjAVperG'] * 16
        df['AdjAVper16G'] = df['AdjAVper16G'].round(decimals=2)


        ##### MOVE IN THE PREDICTION COLUMNS NEEDED here
        # Add the playing year for each player
        df["Yr"] = df["Yr"].astype(int)
        df["CareerYr"] = df.groupby(["PlayerKey"])["Yr"].rank(ascending=True, method='dense')
        # Create the 3 year avg AV column
        df.sort_values(by=["Yr","PlayerKey"], ascending=[True, True], inplace=True)
        df['3YrAvgAV']=df.groupby('Player').AdjAV.apply(lambda x : x.shift().rolling(3,min_periods=1).mean().fillna(x))
        # Capture the player's prior season's AV which is what will be used for forward prediction
        df.sort_values(by=["PlayerKey","Yr"], ascending=[True, True], inplace=True)
        df['PriorYr16GAV'] = df.loc[df['PlayerKey'].shift(-1)==df['PlayerKey'], 'AVper16G']
        df['PriorYr16GAV'] = df['PriorYr16GAV'].shift()
        df['PriorYrAdj16GAV'] = df.loc[df['PlayerKey'].shift(-1)==df['PlayerKey'], 'AdjAVper16G']
        df['PriorYrAdj16GAV'] = df['PriorYrAdj16GAV'].shift()
        df.PriorYr16GAV.fillna(value=0, inplace=True)
        df.PriorYrAdj16GAV.fillna(value=0, inplace=True)
        # Need prior year games and games started to calculate prior year full year AV based on games started (not games played as above)
        df['PriorYrG'] = df.loc[df['PlayerKey'].shift(-1)==df['PlayerKey'], 'G']
        df['PriorYrG'] = df['PriorYrG'].shift()
        df['PriorYrGS'] = df.loc[df['PlayerKey'].shift(-1)==df['PlayerKey'], 'GS']
        df['PriorYrGS'] = df['PriorYrGS'].shift()
        df.PriorYrG.fillna(value=0, inplace=True)
        df.PriorYrGS.fillna(value=0, inplace=True)

        # Create unique player year column which will be used to match to starter df at end
        df['UniqueKey'] = df['Yr'].astype(str) + "-" + df['Tm'] + "-" + df['Player']

        # rearrange columns after adding the abbreviation
        df = df[["Yr", "Player", "PlayerKey", "Age", "Tm", "Pos", "PosGroup", "PositionGroup", "G", "GS", "AV", "AVperG", "AVper16G",
                "AdjAV", "AdjAVperG", "AdjAVper16G", "CareerYr", "3YrAvgAV", "PriorYr16GAV", "PriorYrAdj16GAV", "PriorYrG", "PriorYrGS", "UniqueKey"]]
        df = df.sort_values(by=['Tm','Yr','Player'], ascending=[True, False, True])

        # write this team and season of player data to player csv file
        if not os.path.isfile('player-data-inprocess.csv'):
            df.to_csv('player-data-inprocess.csv', header='column_names')
        else: # else it exists so append without writing the header
            df.to_csv('player-data-inprocess.csv', mode='a', header=False)

        # write this team and season of player data to player csv file
        if not os.path.isfile('player-starter-data.csv'):
            starterdf.to_csv('player-starter-data.csv', header='column_names')
        else: # else it exists so append without writing the header
            starterdf.to_csv('player-starter-data.csv', mode='a', header=False)

        starterdf = starterdf[0:0]

        # Update flag in teams file to mark that we completed this team and season
        scrapedflag_cell = sheet_obj.cell(row=index,column=23)
        scrapedflag_cell.value = 1;
        wb_obj.save(teamsdatafilename)

    else:
        print("Skipping " + str(year) + " " + team + " - already scraped")

# Set starters after everything finished
df = pd.read_csv("player-data-inprocess.csv")
print(df)
df = df[(df.Yr <= 2020)]

starterdf = pd.read_csv("player-starter-data.csv")
print(starterdf)

#df = pd.merge(df, starterdf, how='outer', on=['UniqueKey']) #left_on=['Yr','OrigTm','Player'], right_on = ['Yr','OrigTm','Player'])
df = pd.merge(df, starterdf, how='outer', on=['UniqueKey'])
df = df[["Yr", "Player", "PlayerKey", "Age", "Tm", "Pos", "PosGroup", "PositionGroup", "G", "GS", "AV", "AVperG", "AVper16G",
        "AdjAV", "AdjAVperG", "AdjAVper16G", "CareerYr", "3YrAvgAV", "PriorYr16GAV", "PriorYrAdj16GAV", "PriorYrG", "PriorYrGS", "Starter"]]
df = df.sort_values(by=['Tm','Yr','Player'], ascending=[True, False, True])
# Write final CSV of all player and starter data
df.to_csv('player-data-out.csv', header='column_names')

print("\nFINAL...")
print(str(totalcount) + " players scraped")
print(str(skipped_count) + " players skipped")

# Display the positons scraped that are not in the set list above - these will need to be handled as PFR
# adds any variation of positions on the season records. Add to the posdf dataframe above and re-run to correct any
prefixes = ['Off','Def','ST']
tdf = df[~df.PositionGroup.str.startswith(tuple(prefixes))]
missingpos = len(tdf)
if missingpos > 0:
    print("\nPositions not found")
    print(tdf)
